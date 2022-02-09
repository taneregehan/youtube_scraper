from youtube_comment_scraper_python import *
import pandas as pd
import time
import tkinter
from googletrans import *
from textblob import TextBlob
from openpyxl import load_workbook
from matplotlib import pyplot as plt
from tkinter import messagebox

def youtube_comment(): #Youtubedeki Yorumların Çekilmesi
	youtube.open("https://www.youtube.com/watch?v=FKA28nuw93Y")

	data = []
	currentpagesource = youtube.get_page_source()
	lastpagesource = ''

	while (True):
		if (lastpagesource == currentpagesource):
			break

		lastpagesource = currentpagesource
		response = youtube.video_comments()

		for c in response['body']:
			data.append(c)

		youtube.scroll()
		currentpagesource = youtube.get_page_source()

	df = pd.DataFrame(data)

	df = df.replace('\n', ' ', regex=True)

	df = df[['Comment']].drop_duplicates(keep="first")

	df.to_csv('data.txt', index=False)

wb = load_workbook('C:\\Users\\taner\\OneDrive\\Masaüstü\\Gorus.xlsx')
sheet = wb.active
dizi = []
dosya = open("data.txt", "r", encoding="utf8")
dizi = dosya.readlines()
positive_sentences = 0
negative_sentences = 0
neutral_sentences = 0
translate_sentences = []
translate_sentences2 = []
ldn_words = []
translator = Translator()
def data_read(): #Yorumların Çıktısı
	for i in dizi:
		print("Yorum: "+i)
def translate ():#Yorumların Çevirilmesi
	for i in dizi:
		translate = translator.translate(i, dest='en') #Çeviri Yapıyoruz
		time.sleep(0.3)
		translate_sentences.append(str(translate.text))
	for i in translate_sentences:
		print("Çevirilmiş Cümle: "+i)

def opinion_analysis (): #Yorumların Analizi
	global positive_sentences
	global negative_sentences
	global neutral_sentences
	for i in translate_sentences:
		text1 = TextBlob(i) #TextBlob Kütüphanesiyle yapmamızın sebebi duyguların analiz edilmesini bulmamıza yardımcı olması
		translate_sentences2.append(text1)
	for i in range(len(translate_sentences2)):
		if(translate_sentences2[i].polarity > 0):
			ldn_words.append("Beğenildi :)")
			positive_sentences = positive_sentences + 1
		elif(translate_sentences2[i].polarity < 0):
			ldn_words.append("Olumsuz :(")
			negative_sentences = negative_sentences + 1
		else:
			ldn_words.append("Eşit :|")
			neutral_sentences = neutral_sentences + 1
	print(ldn_words)
	print("Yorumun Beğeni Sayısı: "+str(positive_sentences))
	print("Yorumun Beğenilmeme Sayısı: "+str(negative_sentences))
	print("Yoruma Tarafsız Kalma Sayısı:  "+str(neutral_sentences))
def excel_write(): #Verilerin Excele Yazılması
	row = 2
	sheet['A1'] = "Görüşler"
	for i in range (len(ldn_words)):
		sheet.cell(row=row, column=1, value=ldn_words[i]) #Excele kaydetme
		row+=1
	wb.save(('C:\\Users\\taner\\OneDrive\\Masaüstü\\Gorus.xlsx'))
def analysis_graphic(): #Dijital Olarak Grafiğe Dönüştürülmesi
	pieces =[positive_sentences, negative_sentences, neutral_sentences] #Fikirler
	opinions = ["Beğenen","Beğenmeyen","Kafası Karışık"] #Grafikteki Bölümler
	color = ["g","r","y"]
	plt.pie(pieces,
			labels = opinions,
			colors = color,
			autopct = '%1.1f%%')
	plt.title = "Youtube Video Yorum Analizi"
	plt.show()

def future_comment_analysis(): # Analiz Sonucunun mesaj Kutusunda Gösterilmesi
	if(positive_sentences >= negative_sentences and positive_sentences >= neutral_sentences):
		tkinter.messagebox.showinfo("ANALİZ","GELEN İZLEYİCİ YÜKSEK İHTİMALLE VİDEOYU BEĞENECEK")

	elif(negative_sentences >= positive_sentences and negative_sentences >= neutral_sentences):
		tkinter.messagebox.showinfo("ANALİZ","GELEN İZLEYİCİ YÜKSEK İHTİMALLE VİDEOYU BEĞENMEYECEK")

	else:
		tkinter.messagebox.showinfo("ANALİZ","GELEN İZLEYİCİ VİDEOYA YORUMSUZ ")

youtube_comment()
data_read()
translate()
opinion_analysis()
excel_write()
analysis_graphic()
future_comment_analysis()
wb.close()
